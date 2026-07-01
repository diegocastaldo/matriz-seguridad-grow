"""
motor.py — Matriz de Seguridad Grow
Motor de procesamiento: carga, resolución y exportación.
"""

import pandas as pd
import datetime
from io import BytesIO


# ── CONSTANTS ─────────────────────────────────────────────────────────────────

SAP_BR_DIRECT = [
    "SAP_BR_AA_ACCOUNTANT", "SAP_BR_ANALYTICS_SPECIALIST",
    "SAP_BR_CASH_MANAGER", "SAP_BR_CASH_SPECIALIST",
    "SAP_BR_GL_ACCOUNTANT", "SAP_BR_INVENTORY_MANAGER",
    "SAP_BR_PRICING_SPECIALIST", "SAP_BR_PRODN_OPTR_DISC",
    "SAP_BR_PRODN_PLNR", "SAP_BR_PRODN_SUPERVISOR_DISC",
    "SAP_BR_PURCHASING_MANAGER", "SAP_BR_RECEIVING_SPECIALIST",
    "SAP_BR_SHIPPING_SPECIALIST",
]

MAX_ROLES_PER_APP = 5   # specificity filter
SEP = " — "             # separator in dropdown: "F1443A — Manage Cost Centers"


# ── FILE LOADERS ──────────────────────────────────────────────────────────────

def _parse_sap_unconverted(file) -> pd.DataFrame:
    """Parse SAP SE16N 'Unconverted' pipe-delimited .txt export."""
    if hasattr(file, 'read'):
        raw = file.read()
        if hasattr(file, 'seek'):
            file.seek(0)
    else:
        with open(file, 'rb') as f:
            raw = f.read()

    for enc in ('utf-8', 'latin-1', 'cp1252', 'utf-16'):
        try:
            content = raw.decode(enc)
            break
        except Exception:
            continue

    lines = content.splitlines()
    hdr_idx = next((i for i, l in enumerate(lines)
                    if l.startswith('|') and 'Client ID' in l), None)
    if hdr_idx is None:
        raise ValueError("No se encontró la línea de encabezado en el .txt. "
                         "Verificar que sea export SE16N 'Unconverted'.")

    col_names = [c.strip() for c in lines[hdr_idx].split('|')]
    rows = []
    for line in lines[hdr_idx + 1:]:
        if not line.startswith('|') or line.startswith('|---'):
            continue
        vals = [c.strip() for c in line.split('|')]
        while len(vals) < len(col_names):
            vals.append('')
        rows.append(vals[:len(col_names)])

    return pd.DataFrame(rows, columns=col_names)


def _is_txt(file) -> bool:
    return getattr(file, 'name', '').lower().endswith('.txt')


def load_agr_users(file) -> pd.DataFrame:
    """Load AGR_USERS (.xlsx or .txt). Returns df with User, Role, End."""
    if _is_txt(file):
        df = _parse_sap_unconverted(file)
        col_map = {'Role': 'Role', 'User': 'User',
                   'Start date': 'Start', 'End date': 'End'}
    else:
        df = pd.read_excel(file)
        col_map = {}
        for c in df.columns:
            cl = c.lower()
            if "user" in cl and ("name" in cl or "id" in cl):
                col_map[c] = "User"
            elif cl in ("uname",):
                col_map[c] = "User"
            elif "role" in cl and "id" not in cl:
                col_map[c] = "Role"
            elif cl in ("agr_name",):
                col_map[c] = "Role"
            elif "start" in cl or "from" in cl or cl in ("from_dat",):
                col_map[c] = "Start"
            elif "end" in cl or "to" in cl or cl in ("to_dat",):
                col_map[c] = "End"

    df = df.rename(columns=col_map)
    df = df.loc[:, ~df.columns.duplicated(keep="first")]

    if "User" not in df.columns or "Role" not in df.columns:
        raise ValueError("No se encontraron columnas Usuario/Rol en AGR_USERS.")

    if "End" in df.columns:
        df["End"] = pd.to_datetime(df["End"], dayfirst=True, errors="coerce").dt.date

    keep = ["User", "Role"] + (["End"] if "End" in df.columns else [])
    df = df[[c for c in keep if c in df.columns]].copy()
    df = df[df["User"].notna() & (df["User"] != "") & (df["User"] != "X")]
    return df.drop_duplicates()


def load_agr_tcodes(file) -> pd.DataFrame:
    """Load AGR_TCODES (.xlsx or .txt). Returns df with Role, TCode."""
    if _is_txt(file):
        df = _parse_sap_unconverted(file)
        col_map = {'Role': 'Role', 'ReportTyp': 'ReportTyp',
                   'Extended name': 'TCode'}
    else:
        df = pd.read_excel(file)
        col_map = {}
        for c in df.columns:
            cl = c.lower()
            if cl in ("agr_name", "role"):
                col_map[c] = "Role"
            elif cl in ("tcode", "extended name", "transaction", "transaction code"):
                col_map[c] = "TCode"
            elif "report" in cl and "typ" in cl:
                col_map[c] = "ReportTyp"

    df = df.rename(columns=col_map)
    if "ReportTyp" in df.columns:
        df = df[df["ReportTyp"].str.strip() == "TR"]
    if "Role" not in df.columns or "TCode" not in df.columns:
        raise ValueError("No se encontraron columnas Rol/TCode en AGR_TCODES.")
    df = df[df["Role"].notna() & (df["Role"] != "") & (df["Role"] != "X")]
    df = df[df["TCode"].notna() & (df["TCode"] != "")]
    return df[["Role", "TCode"]].drop_duplicates()


def load_iam_apps(file) -> pd.DataFrame:
    """Load IAM Information System export (Business Role Template - IAM App)."""
    df = pd.read_excel(file)
    col_map = {}
    for c in df.columns:
        cl = c.lower()
        if "business role template id" in cl:
            col_map[c] = "BRT_ID"
        elif "business role template" in cl and "id" not in cl:
            col_map[c] = "BRT_Name"
        elif "business catalog" in cl:
            col_map[c] = "Catalog_ID"
        elif cl == "iam app id":
            col_map[c] = "IAM_App_ID"
        elif cl == "iam app":
            col_map[c] = "IAM_App_Name"
        elif "transaction code" in cl:
            col_map[c] = "Transaction_Code"
    df = df.rename(columns=col_map)
    if "BRT_ID" not in df.columns or "Transaction_Code" not in df.columns:
        raise ValueError("Columnas no encontradas en IAM Apps. "
                         "Verificar que sea export 'Business Role Template - IAM App'.")
    return df


def load_st03n(files: list) -> pd.DataFrame:
    """Load one or more ST03N exports. Returns df with TCode, Executions."""
    dfs = []
    for f in files:
        df = pd.read_excel(f)
        tcode_col = exec_col = None
        for c in df.columns:
            cl = c.lower()
            if tcode_col is None and any(k in cl for k in
                    ("report", "transaction", "program")):
                tcode_col = c
            if exec_col is None and any(k in cl for k in
                    ("step", "exec", "count", "number", "dialog")):
                exec_col = c
        if tcode_col and exec_col:
            tmp = df[[tcode_col, exec_col]].copy()
            tmp.columns = ["TCode", "Executions"]
            tmp["Executions"] = pd.to_numeric(
                tmp["Executions"], errors="coerce").fillna(0)
            dfs.append(tmp)
    if not dfs:
        return pd.DataFrame(columns=["TCode", "Executions"])
    combined = pd.concat(dfs, ignore_index=True)
    combined["Executions"] = pd.to_numeric(
        combined["Executions"], errors="coerce").fillna(0)
    return combined.groupby("TCode")["Executions"].sum().reset_index()


def load_equivalencias_completadas(file) -> pd.DataFrame:
    """
    Load a completed equivalencias file.
    Handles our generated format (header at row 3) and any generic .xlsx.
    App_Grow_ID may contain 'F1443A — App Name' format; we strip to just the ID.
    """
    def read_raw(sheet_name):
        """Read a sheet with no header row interpretation."""
        kw = {"header": None}
        if sheet_name:
            kw["sheet_name"] = sheet_name
        if hasattr(file, 'seek'):
            file.seek(0)
        try:
            return pd.read_excel(file, **kw)
        except Exception:
            return None

    def find_header_row(df_raw):
        """Find the 0-indexed row that contains a TCode column header."""
        for i, row in df_raw.iterrows():
            vals = [str(v).lower().strip() for v in row.values]
            if any("tcode" in v and len(v) < 30 for v in vals):
                return i
        return None

    df = None
    for sheet_name in ("Equivalencias a Completar", "Equivalencias", None):
        raw = read_raw(sheet_name)
        if raw is None:
            continue
        hdr_row = find_header_row(raw)
        if hdr_row is not None:
            kw = {"header": int(hdr_row)}
            if sheet_name:
                kw["sheet_name"] = sheet_name
            if hasattr(file, 'seek'):
                file.seek(0)
            df = pd.read_excel(file, **kw)
            break

    if df is None:
        raise ValueError("No se pudo leer el archivo de equivalencias.")

    # Normalise column names
    col_map = {}
    for c in df.columns:
        cl = str(c).lower().strip()
        if "tcode" in cl and len(cl) < 30:
            col_map[c] = "TCode_GUI"
        elif "app_grow_id" in cl or "app grow id" in cl or "equivalencia" in cl:
            col_map[c] = "App_Grow_ID"
    df = df.rename(columns=col_map)

    if "TCode_GUI" not in df.columns or "App_Grow_ID" not in df.columns:
        raise ValueError(
            "El archivo de equivalencias debe tener columnas 'TCode_GUI' y 'App_Grow_ID'. "
            f"Columnas encontradas: {list(df.columns)}")

    def extract_id(val):
        val = str(val).strip()
        if val in ("nan", "None", ""):
            return ""
        if SEP in val:
            return val.split(SEP)[0].strip()
        return val

    df["App_Grow_ID"] = df["App_Grow_ID"].apply(extract_id)
    df = df[df["TCode_GUI"].notna() & (df["TCode_GUI"].astype(str).str.strip() != "")]
    return df[["TCode_GUI", "App_Grow_ID"]].reset_index(drop=True)


# ── HELPERS ───────────────────────────────────────────────────────────────────

def build_active_user_roles(df_users: pd.DataFrame) -> pd.DataFrame:
    today = datetime.date.today()
    if "End" in df_users.columns:
        active = df_users[df_users["End"].isna() | (df_users["End"] >= today)]
    else:
        active = df_users.copy()
    return active[["User", "Role"]].drop_duplicates()


def build_iam_map(df_iam: pd.DataFrame) -> dict:
    """TCode -> set of Grow Role Template IDs, specificity filter ≤5."""
    df = df_iam.loc[:, ~df_iam.columns.duplicated()]
    cnt = df.groupby("Transaction_Code")["BRT_ID"].nunique()
    specific = df[df["Transaction_Code"].map(cnt) <= MAX_ROLES_PER_APP]
    return (specific.groupby("Transaction_Code")["BRT_ID"]
            .apply(set).to_dict())


def _iam_lookups(df_iam: pd.DataFrame):
    """Return (iam_tcodes set, tc->app_name dict, app_name->tc dict)."""
    df = df_iam.loc[:, ~df_iam.columns.duplicated()]
    iam_tcodes = set(df["Transaction_Code"].dropna().str.strip())
    dedup = df.drop_duplicates("Transaction_Code")
    tc_to_name = dict(zip(dedup["Transaction_Code"].values,
                          dedup["IAM_App_Name"].values))
    name_to_tc = {}
    for _, row in df.iterrows():
        name = str(row.get("IAM_App_Name", "")).strip().lower()
        tc = str(row.get("Transaction_Code", "")).strip()
        if name and tc and name not in name_to_tc:
            name_to_tc[name] = tc
    return iam_tcodes, tc_to_name, name_to_tc


# ── ETAPA 1: GENERAR EQUIVALENCIAS ────────────────────────────────────────────

def generate_equivalencias_excel(
    df_users: pd.DataFrame,
    df_tcodes: pd.DataFrame,
    df_iam: pd.DataFrame,
    df_st03n: pd.DataFrame,
    cliente: str = "",
    proyecto: str = "",
) -> bytes:
    """
    Generate the 'Equivalencias a Completar' Excel for consultants.
    - All active TCodes (used in ST03N + assigned to active users)
    - Auto-resolved ones come pre-filled
    - App_Grow_ID column has a dropdown: "F1443A — Manage Cost Centers"
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, Protection)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.styles.numbers import FORMAT_DATE_DATETIME

    # ── Active TCodes with stats ───────────────────────────────────────────
    df_ur = build_active_user_roles(df_users)
    active_roles = set(df_ur["Role"].dropna().unique())
    df_rt = df_tcodes[df_tcodes["Role"].isin(active_roles)].copy()

    user_tcode = df_ur.merge(df_rt, on="Role", how="inner")
    tcode_users = user_tcode.groupby("TCode")["User"].nunique().to_dict()
    tcode_roles = df_rt.groupby("TCode")["Role"].apply(
        lambda x: ", ".join(sorted(set(x)))).to_dict()

    exec_map = {}
    used_tcodes = None
    if df_st03n is not None and len(df_st03n) > 0:
        exec_map = {str(tc): float(ex)
                    for tc, ex in zip(df_st03n["TCode"].values,
                                      df_st03n["Executions"].values)
                    if not hasattr(ex, '__len__')}
        used_tcodes = set(df_st03n["TCode"].str.strip().unique())

    all_tcodes = set(df_rt["TCode"].dropna().str.strip().unique())
    if used_tcodes:
        active_tcodes = all_tcodes & used_tcodes
    else:
        active_tcodes = all_tcodes

    # ── IAM lookups ────────────────────────────────────────────────────────
    iam_tcodes, tc_to_name, _ = _iam_lookups(df_iam)

    # ── Build dropdown list ────────────────────────────────────────────────
    # Format: "AppID — App Name", sorted alphabetically by name
    df_dd = (df_iam.loc[:, ~df_iam.columns.duplicated()]
             .drop_duplicates(["Transaction_Code", "IAM_App_Name"])
             [["Transaction_Code", "IAM_App_Name"]]
             .dropna())
    df_dd = df_dd[df_dd["Transaction_Code"].str.strip() != ""]
    df_dd["dropdown"] = (df_dd["Transaction_Code"].str.strip()
                         + SEP
                         + df_dd["IAM_App_Name"].str.strip())
    dropdown_values = sorted(df_dd["dropdown"].unique().tolist(),
                             key=lambda x: x.split(SEP)[-1].lower())

    # ── Build rows ─────────────────────────────────────────────────────────
    rows = []
    for tc in sorted(active_tcodes):
        tc = str(tc).strip()
        execs = int(exec_map.get(tc, 0))
        users = int(tcode_users.get(tc, 0))

        if tc in iam_tcodes:
            app_name = str(tc_to_name.get(tc, "")).strip()
            prefill = f"{tc}{SEP}{app_name}" if app_name else tc
            estado = "✅ Propuesto (automático)"
        else:
            prefill = ""
            estado = "⬅ Completar"

        rows.append({
            "TCode_GUI": tc,
            "App_Grow_ID": prefill,
            "Ejecuciones_3m": execs,
            "Usuarios_Afectados": users,
            "Estado": estado,
        })

    # Sort: pending first by executions desc, then auto by executions desc
    df_rows = pd.DataFrame(rows)
    df_rows["_ord"] = df_rows["Estado"].apply(
        lambda x: 0 if "Completar" in x else 1)
    df_rows = (df_rows
               .sort_values(["_ord", "Ejecuciones_3m"],
                            ascending=[True, False])
               .drop("_ord", axis=1)
               .reset_index(drop=True))

    # ── Build Excel ────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    NAVY   = "1F3864"
    BLUE   = "2E75B6"
    GREEN  = "E2EFDA"
    YELLOW = "FFFACD"
    ORANGE = "FFF2CC"
    WHITE  = "FFFFFF"
    GREY   = "F5F5F5"

    def hfill(c):
        return PatternFill("solid", start_color=c, fgColor=c)

    def bdr(color="D0D7E3"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def widths(ws, wl):
        for i, w in enumerate(wl, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Main sheet (created FIRST so it's the default sheet) ─────────────
    ws = wb.create_sheet("Equivalencias a Completar")

    # ── Hidden reference sheet with dropdown values (created after main) ───
    ws_ref = wb.create_sheet("_IAM_Ref")
    ws_ref.sheet_state = "hidden"
    ws_ref.cell(row=1, column=1, value="Equivalencia Grow (ID — Nombre)")
    for i, val in enumerate(dropdown_values, 2):
        ws_ref.cell(row=i, column=1, value=val)

    # Named range pointing to dropdown list
    n = len(dropdown_values)
    ref_formula = f"'_IAM_Ref'!$A$2:$A${n+1}"
    wb.defined_names["_IAM_Apps_List"] = DefinedName(
        "_IAM_Apps_List", attr_text=ref_formula)
    ws.sheet_properties.tabColor = "2E75B6"
    ws.freeze_panes = "A4"

    # Row 1: Title
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    meta = f" {proyecto} — {cliente}" if (proyecto or cliente) else ""
    c.value = f"Equivalencias a Completar{meta}"
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = hfill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2: Subtitle / instructions
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:F2")
    c2 = ws["A2"]
    c2.value = (f"Generado: {datetime.date.today().strftime('%d/%m/%Y')}  |  "
                f"{len(df_rows)} TCodes activos  |  "
                f"Completar columnas 'App_Grow_ID' (desplegable) y 'Consultor'")
    c2.font = Font(name="Arial", italic=True, size=9, color="555555")
    c2.fill = hfill("EEF3FA")
    c2.alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Headers
    ws.row_dimensions[3].height = 22
    headers = [
        ("TCode_GUI",          "TCode GUI (Private)",        14, NAVY,     True),
        ("App_Grow_ID",        "App_Grow_ID  ✏ Completar ▼", 52, "0070B8", True),
        ("Consultor",          "Consultor  ✏",               20, "5B4DA0", True),
        ("Ejecuciones_3m",     "Ejecuciones (3m)",           18, NAVY,     True),
        ("Usuarios_Afectados", "Usuarios Afectados",         18, NAVY,     True),
        ("Estado",             "Estado",                     24, NAVY,     True),
    ]
    for ci, (_, label, width, bg, bold) in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=label)
        c.font = Font(name="Arial", bold=bold, color="FFFFFF", size=10)
        c.fill = hfill(bg)
        c.border = bdr()
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    n_data = len(df_rows)
    for i, (_, row) in enumerate(df_rows.iterrows()):
        r = 4 + i
        ws.row_dimensions[r].height = 16

        is_auto = "automático" in str(row["Estado"])

        vals = [row["TCode_GUI"], row["App_Grow_ID"], "",   # Consultor empty
                row["Ejecuciones_3m"], row["Usuarios_Afectados"],
                row["Estado"]]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = Font(name="Arial", size=9,
                          color="375623" if is_auto else "333333")
            c.border = bdr()
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if ci in (4, 5) else "left")

            # Col A: TCode — locked
            if ci == 1:
                c.fill = hfill(GREY)
                c.protection = Protection(locked=True)
            # Col B: App_Grow_ID — editable, dropdown
            elif ci == 2:
                c.fill = hfill(GREEN if is_auto else YELLOW)
                c.font = Font(name="Arial", size=9,
                              bold=not is_auto,
                              color="375623" if is_auto else "7E3000")
                c.protection = Protection(locked=False)
            # Col C: Consultor — editable, light purple tint
            elif ci == 3:
                c.fill = hfill("F3F0FA")
                c.protection = Protection(locked=False)
            # Cols D, E, F: read-only stats & status
            else:
                c.fill = hfill(GREY if i % 2 == 0 else WHITE)
                c.protection = Protection(locked=True)

    # Dropdown on column B (App_Grow_ID)
    dv = DataValidation(
        type="list",
        formula1="_IAM_Apps_List",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valor no válido",
        error="Seleccioná una opción del desplegable o dejá vacío.",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"B4:B{4 + n_data - 1}"

    # Legend row below data
    legend_row = 4 + n_data + 1
    ws.row_dimensions[legend_row].height = 16
    ws.merge_cells(f"A{legend_row}:F{legend_row}")
    lc = ws.cell(row=legend_row, column=1,
                 value="✅ Verde = resuelto automáticamente  |  "
                       "🟡 Amarillo = completar App_Grow_ID  |  "
                       "Columna 'Consultor': nombre del consultor que validó la equivalencia")
    lc.font = Font(name="Arial", size=8, italic=True, color="666666")
    lc.fill = hfill("F0F4FA")
    lc.alignment = Alignment(horizontal="left", vertical="center")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── ETAPA 2: CALCULAR ROLES ───────────────────────────────────────────────────

def run_motor(
    df_users: pd.DataFrame,
    df_tcodes: pd.DataFrame,
    df_iam: pd.DataFrame,
    df_equiv: pd.DataFrame,
    df_st03n: pd.DataFrame | None = None,
    progress_cb=None,
) -> dict:
    """
    Main calculation. Returns dict with df_output, df_motor,
    df_mapeo, df_gap, stats.
    """
    def progress(msg, pct):
        if progress_cb:
            progress_cb(msg, pct)

    progress("Filtrando usuarios activos...", 5)
    df_ur = build_active_user_roles(df_users)

    progress("Construyendo mapa IAM...", 15)
    iam_map = build_iam_map(df_iam)          # ≤5 roles: for role assignment
    iam_tcodes, tc_to_name, name_to_tc = _iam_lookups(df_iam)
    # iam_tcodes = ALL TCodes in IAM regardless of role count (for "mapped?" check)
    df_iam_d = df_iam.loc[:, ~df_iam.columns.duplicated()]
    iam_role_name = dict(zip(df_iam_d["BRT_ID"].values,
                              df_iam_d["BRT_Name"].values))

    progress("Procesando equivalencias...", 25)
    eq_lookup = {}
    for _, r in df_equiv.iterrows():
        tc = str(r.get("TCode_GUI", "")).strip()
        raw = str(r.get("App_Grow_ID", "")).strip()
        # Strip " — Name" suffix if present
        app_id = raw.split(SEP)[0].strip() if SEP in raw else raw
        if tc and app_id and app_id not in ("nan", "None", ""):
            eq_lookup[tc] = app_id

    used_tcodes = None
    if df_st03n is not None and len(df_st03n) > 0:
        used_tcodes = set(df_st03n["TCode"].str.strip().unique())

    active_roles = set(df_ur["Role"].dropna().unique())
    df_rt = df_tcodes[df_tcodes["Role"].isin(active_roles)].copy()

    progress("Calculando asignaciones de roles...", 40)

    user_grow_roles: dict[str, set] = {}
    user_tcode_status: dict[str, dict] = {}

    for _, ur_row in df_ur.iterrows():
        user = ur_row["User"]
        role = ur_row["Role"]

        if user not in user_grow_roles:
            user_grow_roles[user] = set()
            user_tcode_status[user] = {}

        if role in SAP_BR_DIRECT:
            user_grow_roles[user].add(role)

        role_tcodes = df_rt[df_rt["Role"] == role]["TCode"].unique()
        for tc in role_tcodes:
            tc = str(tc).strip()
            grow_roles = set()

            if tc in iam_map:
                # Direct IAM match, ≤5 roles → assign those roles
                grow_roles = iam_map[tc]
            elif tc in eq_lookup and eq_lookup[tc] in iam_map:
                # Via equivalencias, ≤5 roles → assign those roles
                grow_roles = iam_map[eq_lookup[tc]]

            if grow_roles:
                user_grow_roles[user].update(grow_roles)
                user_tcode_status[user][tc] = "mapped"
            elif tc in iam_tcodes:
                # TCode exists in IAM but app appears in >5 roles (e.g. ME23N, FB03).
                # Mark as mapped — access comes through other roles already assigned.
                # Do NOT add to gap.
                user_tcode_status[user][tc] = "mapped"
            elif tc in eq_lookup and eq_lookup[tc] in iam_tcodes:
                # Via equivalencias, app exists in IAM but >5 roles
                user_tcode_status[user][tc] = "mapped"
            else:
                if tc not in user_tcode_status[user]:
                    user_tcode_status[user][tc] = "gap"

    progress("Calculando gaps...", 65)

    no_equiv_tcodes: set = set()

    progress("Construyendo OUTPUT...", 75)
    output_rows, motor_rows, gap_rows = [], [], []

    for user in sorted(user_grow_roles.keys()):
        roles = sorted(user_grow_roles[user])
        pending = {
            tc for tc, st in user_tcode_status[user].items()
            if st == "gap"
            and tc not in no_equiv_tcodes
            and (used_tcodes is None or tc in used_tcodes)
        }
        output_rows.append({
            "Usuario": user,
            "Roles Grow a Asignar": ", ".join(roles),
            "Cant. Roles": len(roles),
            "TCodes Pendientes": ", ".join(sorted(pending)) if pending else "",
            "Cant. Pendientes": len(pending),
            "Estado": "✅ Completo" if not pending else f"⚠️ {len(pending)} pendientes",
        })
        for r in roles:
            motor_rows.append({
                "Usuario": user,
                "Grow Role Template ID": r,
                "Grow Role Template": iam_role_name.get(r, r),
            })
        for tc in pending:
            gap_rows.append({"Usuario": user, "TCode": tc})

    progress("Construyendo Mapeo Detalle...", 85)
    mapeo_rows = []
    for _, ur_row in df_ur.iterrows():
        user, role = ur_row["User"], ur_row["Role"]
        for tc in df_rt[df_rt["Role"] == role]["TCode"].unique():
            tc = str(tc).strip()
            if used_tcodes and tc not in used_tcodes:
                continue
            if user_tcode_status.get(user, {}).get(tc) != "mapped":
                continue
            app_id = (tc if tc in iam_map
                      else eq_lookup.get(tc, ""))
            app_name = str(tc_to_name.get(app_id, ""))
            for gr in iam_map.get(app_id, set()):
                mapeo_rows.append({
                    "Usuario": user, "TCode Private": tc, "Rol Private": role,
                    "App Grow ID": app_id, "App Grow Nombre": app_name,
                    "Rol Grow ID": gr,
                    "Rol Grow Nombre": iam_role_name.get(gr, gr),
                })

    df_mapeo = (pd.DataFrame(mapeo_rows)
                .drop_duplicates(subset=["Usuario", "TCode Private", "Rol Grow ID"])
                if mapeo_rows else pd.DataFrame())

    df_gap_detail = (pd.DataFrame(gap_rows) if gap_rows
                     else pd.DataFrame(columns=["Usuario", "TCode"]))
    if len(df_gap_detail) > 0:
        df_gap_summ = (df_gap_detail.groupby("TCode")
                       .agg(Usuarios_Afectados=("Usuario", "nunique"))
                       .reset_index())
        if df_st03n is not None:
            exec_s = df_st03n.groupby("TCode")["Executions"].sum()
            df_gap_summ["Ejecuciones_3m"] = (
                df_gap_summ["TCode"].map(exec_s).fillna(0).astype(int))
        else:
            df_gap_summ["Ejecuciones_3m"] = 0
        df_gap_summ = df_gap_summ.sort_values("Ejecuciones_3m", ascending=False)
    else:
        df_gap_summ = pd.DataFrame(
            columns=["TCode", "Usuarios_Afectados", "Ejecuciones_3m"])

    progress("Finalizando...", 95)

    df_output = pd.DataFrame(output_rows)
    df_motor = (pd.DataFrame(motor_rows).drop_duplicates()
                if motor_rows else pd.DataFrame())

    stats = {
        "total_users": len(df_output),
        "complete_users": int((df_output["Cant. Pendientes"] == 0).sum())
                          if len(df_output) else 0,
        "total_roles_assigned": (int(df_motor["Grow Role Template ID"].nunique())
                                 if len(df_motor) else 0),
        "gap_tcodes": int(df_gap_summ["TCode"].nunique()) if len(df_gap_summ) else 0,
        "pending_equiv": int((df_equiv["App_Grow_ID"].fillna("") == "").sum())
                         if "App_Grow_ID" in df_equiv.columns else 0,
    }
    progress("¡Listo!", 100)

    return {"df_output": df_output, "df_motor": df_motor,
            "df_mapeo": df_mapeo, "df_gap": df_gap_summ, "stats": stats}


# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────

def export_excel(result: dict, df_equiv: pd.DataFrame,
                 cliente: str = "", proyecto: str = "") -> bytes:
    """Generate the full results Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    def hfill(c):
        return PatternFill("solid", start_color=c, fgColor=c)

    def bdr():
        s = Side(style="thin", color="BDD7EE")
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr(ws, row, vals, bg, fg="FFFFFF", h=22):
        ws.row_dimensions[row].height = h
        for ci, v in enumerate(vals, 1):
            cl = ws.cell(row=row, column=ci, value=v)
            cl.font = Font(name="Arial", bold=True, color=fg, size=10)
            cl.fill = hfill(bg)
            cl.border = bdr()
            cl.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

    def write_df(ws, df, start_row):
        for i, (_, row) in enumerate(df.iterrows()):
            r = start_row + i
            ws.row_dimensions[r].height = 15
            bg = "F2F2F2" if i % 2 == 0 else "FFFFFF"
            for ci, val in enumerate(row.values, 1):
                cl = ws.cell(row=r, column=ci,
                             value=(val if not pd.isna(val) else ""))
                cl.font = Font(name="Arial", size=9)
                cl.fill = hfill(bg)
                cl.border = bdr()
                cl.alignment = Alignment(vertical="center")

    def title_row(ws, txt, sub, ncols=6):
        ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill = hfill("1F3864")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 15
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        ws["A2"].value = sub
        ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        return 3

    def widths(ws, wl):
        for i, w in enumerate(wl, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    today = datetime.date.today().strftime("%d/%m/%Y")
    meta = f"{proyecto} — {cliente}  |  " if (proyecto or cliente) else ""

    # OUTPUT
    ws = wb.create_sheet("OUTPUT")
    ws.sheet_properties.tabColor = "00B050"
    s = title_row(ws, "OUTPUT — Roles Grow a asignar por usuario",
                  f"{meta}Generado: {today}  |  "
                  f"{result['stats']['total_users']} usuarios  |  "
                  f"{result['stats']['complete_users']} completos", 5)
    hdr(ws, s, ["Usuario", "Roles Grow a Asignar", "Cant. Roles",
                "TCodes Pendientes de Mapear", "Cant.", "Estado"], "00B050")
    df_o = result["df_output"]
    for i, (_, row) in enumerate(df_o.iterrows()):
        r = s + 1 + i
        est = str(row["Estado"])
        is_complete = "Completo" in est
        bg = "E2EFDA" if is_complete else "FFF2CC"
        fg = "375623" if is_complete else "7E6000"

        vals = [
            row["Usuario"],
            row["Roles Grow a Asignar"],
            row["Cant. Roles"],
            row["TCodes Pendientes"],   # actual list of pending TCodes
            row["Cant. Pendientes"],    # count
            row["Estado"],
        ]
        for ci, val in enumerate(vals, 1):
            cl = ws.cell(row=r, column=ci,
                         value=(val if not pd.isna(val) else ""))
            cl.font = Font(name="Arial", size=9, color=fg)
            # cols 4, 5, 6 get the status colour; others alternate grey/white
            cl.fill = hfill(
                bg if ci in (4, 5, 6) else ("F2F2F2" if i % 2 == 0 else "FFFFFF"))
            cl.border = bdr()
            cl.alignment = Alignment(
                vertical="center",
                horizontal="center" if ci in (3, 5) else "left",
                wrap_text=(ci in (2, 4)))   # wrap roles AND pending TCodes

        # Row height: taller when pending list is long
        pending_str = str(row["TCodes Pendientes"])
        roles_str   = str(row["Roles Grow a Asignar"])
        max_len = max(len(pending_str), len(roles_str))
        ws.row_dimensions[r].height = max(16, min(80, max_len // 6))

    widths(ws, [18, 85, 12, 55, 10, 24])
    ws.freeze_panes = f"A{s+1}"

    # EQUIVALENCIAS (for reference)
    if df_equiv is not None and len(df_equiv) > 0:
        ws = wb.create_sheet("Equivalencias")
        ws.sheet_properties.tabColor = "ED7D31"
        s = title_row(ws, "EQUIVALENCIAS utilizadas en el cálculo",
                      f"{meta}Generado: {today}", 3)
        hdr(ws, s, ["TCode_GUI", "App_Grow_ID", "Estado"], "1F3864")
        for i, (_, row) in enumerate(df_equiv.iterrows()):
            r = s + 1 + i
            for ci, col in enumerate(["TCode_GUI", "App_Grow_ID"], 1):
                val = str(row.get(col, "")) if row.get(col) else ""
                cl = ws.cell(row=r, column=ci, value=val)
                cl.font = Font(name="Arial", size=9)
                cl.fill = hfill("F2F2F2" if i % 2 == 0 else "FFFFFF")
                cl.border = bdr()
        widths(ws, [20, 55, 20])
        ws.freeze_panes = f"A{s+1}"

    # GAP ACTIVO
    df_gap = result["df_gap"]
    if len(df_gap) > 0:
        ws = wb.create_sheet("GAP_Activo")
        ws.sheet_properties.tabColor = "FF0000"
        s = title_row(ws, f"GAP Activo — {len(df_gap)} TCodes sin equivalencia",
                      f"{meta}Completar en el archivo de equivalencias.", 3)
        hdr(ws, s, ["TCode Private", "Usuarios Afectados", "Ejecuciones 3m"],
            "CC0000")
        write_df(ws, df_gap, s + 1)
        widths(ws, [20, 18, 18])
        ws.freeze_panes = f"A{s+1}"

    # MAPEO DETALLE
    if len(result.get("df_mapeo", pd.DataFrame())) > 0:
        ws = wb.create_sheet("Mapeo_Detalle")
        ws.sheet_properties.tabColor = "7030A0"
        cols = ["Usuario", "TCode Private", "Rol Private",
                "App Grow ID", "App Grow Nombre", "Rol Grow ID", "Rol Grow Nombre"]
        s = title_row(ws, f"Mapeo Detalle — {len(result['df_mapeo']):,} filas",
                      f"{meta}Solo TCodes activos.", 7)
        hdr(ws, s, cols, "7030A0")
        write_df(ws, result["df_mapeo"][cols], s + 1)
        widths(ws, [16, 16, 40, 14, 40, 28, 40])
        ws.freeze_panes = f"A{s+1}"

    # MOTOR
    if len(result.get("df_motor", pd.DataFrame())) > 0:
        ws = wb.create_sheet("MOTOR")
        ws.sheet_properties.tabColor = "4472C4"
        s = title_row(ws, f"MOTOR — {len(result['df_motor']):,} pares Usuario × Rol",
                      f"{meta}Calculado automáticamente.", 3)
        hdr(ws, s, ["Usuario", "Grow Role Template ID", "Grow Role Template"],
            "4472C4")
        write_df(ws, result["df_motor"], s + 1)
        widths(ws, [18, 32, 42])
        ws.freeze_panes = f"A{s+1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
